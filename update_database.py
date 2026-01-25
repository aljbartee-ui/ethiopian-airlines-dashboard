#!/usr/bin/env python3
"""
Database Update Script for Ethiopian Airlines Dashboard
Usage: python3 update_database.py <path_to_excel_file>
"""

import sys
import sqlite3
import json
from datetime import datetime
import openpyxl
from pathlib import Path

def create_database_with_sales_data(excel_path, db_path='src/database/app.db'):
    """
    Create a fresh database with sales data from Excel file
    
    Args:
        excel_path: Path to the Excel file with sales data
        db_path: Path where the database should be created
    """
    print(f"📊 Loading Excel file: {excel_path}")
    
    # Load the Excel file
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")
        sys.exit(1)
    
    # Check if 'CA FOP' sheet exists
    if 'CA FOP' not in wb.sheetnames:
        print(f"❌ Sheet 'CA FOP' not found. Available sheets: {wb.sheetnames}")
        sys.exit(1)
    
    ws = wb['CA FOP']
    print(f"✓ Found 'CA FOP' sheet")
    
    # Extract data
    sales_records = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[1]:  # Check if DATE exists
            record = {
                "Tickets ": row[0],
                "DATE": str(row[1]),
                "Ticket Number": row[2],
                "Amount": row[3],
                "Issuing agent": row[4],
                "FOP": row[5],
                "Time": row[6],
                "INCOME": row[7],
                "Day": row[8],
                "TIME 24HRS": str(row[9]) if row[9] else None
            }
            sales_records.append(record)
    
    print(f"✓ Extracted {len(sales_records)} sales records")
    
    # Calculate totals
    total_tickets = sum(r.get('Tickets ', 0) or 0 for r in sales_records)
    total_revenue = sum(r.get('INCOME', 0) or 0 for r in sales_records)
    
    print(f"  - Total Tickets: {total_tickets}")
    print(f"  - Total Revenue: {total_revenue:.3f} KWD")
    
    # Create database
    print(f"\n🗄️  Creating database: {db_path}")
    
    # Ensure directory exists
    Path(db_path).parent.mkdir(parents=True, exist_ok=True)
    
    # Connect and create fresh database
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Drop existing tables
    cursor.execute("DROP TABLE IF EXISTS sales_data")
    cursor.execute("DROP TABLE IF EXISTS flight_load_data")
    cursor.execute("DROP TABLE IF EXISTS manifest_data")
    cursor.execute("DROP TABLE IF EXISTS route_forecast")
    
    print("✓ Dropped old tables")
    
    # Create sales_data table
    cursor.execute('''
        CREATE TABLE sales_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data TEXT NOT NULL,
            upload_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Insert sales data as JSON
    cursor.execute("INSERT INTO sales_data (data) VALUES (?)", 
                   (json.dumps(sales_records),))
    
    print("✓ Created sales_data table")
    
    # Create flight_load_data table
    cursor.execute('''
        CREATE TABLE flight_load_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            flight_number TEXT,
            flight_date DATE,
            capacity INTEGER,
            forecast INTEGER,
            actual INTEGER,
            load_factor REAL,
            upload_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    print("✓ Created flight_load_data table")
    
    # Create manifest_data table
    cursor.execute('''
        CREATE TABLE manifest_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            flight_number TEXT,
            flight_date DATE,
            total_passengers INTEGER,
            routes TEXT,
            upload_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    print("✓ Created manifest_data table")
    
    # Create route_forecast table
    cursor.execute('''
        CREATE TABLE route_forecast (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            forecast_date DATE,
            flight_direction TEXT,
            route_code TEXT,
            passengers INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    print("✓ Created route_forecast table")
    
    # Commit and close
    conn.commit()
    conn.close()
    
    print(f"\n✅ Database created successfully!")
    print(f"\n📋 Summary:")
    print(f"  - Database: {db_path}")
    print(f"  - Sales Records: {len(sales_records)}")
    print(f"  - Total Tickets: {total_tickets}")
    print(f"  - Total Revenue: {total_revenue:.3f} KWD")
    print(f"\n📝 Next steps:")
    print(f"  1. git add {db_path}")
    print(f"  2. git commit -m \"Update sales data: {total_revenue:.0f} KWD, {total_tickets} tickets\"")
    print(f"  3. git push")
    print(f"\n🚀 Render will auto-deploy in 2-3 minutes")

def main():
    if len(sys.argv) != 2:
        print("Usage: python3 update_database.py <path_to_excel_file>")
        print("\nExample:")
        print("  python3 update_database.py /path/to/DashboardsalesCA.xlsx")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    
    if not Path(excel_path).exists():
        print(f"❌ Error: File not found: {excel_path}")
        sys.exit(1)
    
    create_database_with_sales_data(excel_path)

if __name__ == '__main__':
    main()
