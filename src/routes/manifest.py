from flask import Blueprint, render_template, request, jsonify, session
from src.models.user import db
from src.models.manifest import DailyManifest, RouteForecast, AirportMaster
from datetime import datetime, timedelta
import openpyxl
from io import BytesIO
from collections import defaultdict
import re

manifest_bp = Blueprint('manifest', __name__)

def admin_required(f):
    """Decorator to require admin authentication"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return jsonify({'success': False, 'error': 'Admin authentication required'}), 401
        return f(*args, **kwargs)
    return decorated_function

def parse_text_manifest(content):
    """
    Parse text-based manifest file (Ethiopian Airlines format)
    Returns flight info and passenger breakdown by destination
    """
    lines = content.split('\n')
    
    flight_info = {
        'flight_number': None,
        'date': None,
        'origin': None,
        'destination': None,
        'passengers': [],
        'route_breakdown': defaultdict(int),
        'totals': {
            'male': 0,
            'female': 0,
            'child': 0,
            'infant': 0,
            'bags': 0,
            'weight': 0,
            'total': 0
        }
    }
    
    # Parse header info
    for line in lines[:10]:
        # Flight number and date
        if 'FLIGHT:' in line and 'DATE:' in line:
            # Extract flight number (e.g., "ET  621" -> "ET621")
            flight_match = re.search(r'FLIGHT:\s*(\w+)\s*(\d+)', line)
            if flight_match:
                flight_info['flight_number'] = f"{flight_match.group(1)}{flight_match.group(2)}"
            
            # Extract date (e.g., "03JAN26" -> 2026-01-03)
            date_match = re.search(r'DATE:\s*(\d{2})([A-Z]{3})(\d{2})', line)
            if date_match:
                day = date_match.group(1)
                month_str = date_match.group(2)
                year = date_match.group(3)
                
                months = {'JAN': '01', 'FEB': '02', 'MAR': '03', 'APR': '04', 
                          'MAY': '05', 'JUN': '06', 'JUL': '07', 'AUG': '08',
                          'SEP': '09', 'OCT': '10', 'NOV': '11', 'DEC': '12'}
                month = months.get(month_str, '01')
                
                # Assume 20xx for year
                full_year = f"20{year}"
                flight_info['date'] = f"{full_year}-{month}-{day}"
        
        # Origin and destination
        if 'PT.OF EMBARKATION:' in line:
            emb_match = re.search(r'PT\.OF EMBARKATION:\s*(\w+)', line)
            if emb_match:
                flight_info['origin'] = emb_match.group(1)
            
            dest_match = re.search(r'PT\.OF DEST:\s*(\w+)', line)
            if dest_match:
                flight_info['destination'] = dest_match.group(1)
    
    # Parse passenger lines
    # Format: 001 ABDALLA/ABDEL/M./31A/..0/....../....../0712157554673/......./.../ET00348/PZU/....
    passenger_pattern = re.compile(r'^(\d{3})\s+([A-Z]+)/([A-Z\s]+)/([MF])\.?/(\d+[A-Z])?/')
    
    for line in lines:
        line = line.strip()
        if not line or line.startswith('.') or line.startswith('-') or line.startswith('TOTALS'):
            continue
        
        match = passenger_pattern.match(line)
        if match:
            pax_num = match.group(1)
            last_name = match.group(2)
            first_name = match.group(3).strip()
            gender = match.group(4)
            seat = match.group(5) if match.group(5) else ''
            
            # Extract route code from the line
            # For ET620 (inbound ADD->KWI): /ET00348/PZU/ means passenger came FROM PZU (origin)
            # For ET621 (outbound KWI->ADD): /ET00348/PZU/ means passenger going TO PZU (destination)
            route_match = re.search(r'/ET\d+/([A-Z]{3})/', line)
            if route_match:
                route_code = route_match.group(1)
            else:
                # If no connecting flight, use the main destination/origin
                route_code = flight_info['destination'] or flight_info['origin'] or 'ADD'
            
            passenger = {
                'number': pax_num,
                'name': f"{last_name}/{first_name}",
                'gender': gender,
                'seat': seat,
                'route_code': route_code
            }
            flight_info['passengers'].append(passenger)
            flight_info['route_breakdown'][route_code] += 1
            
            # Count by gender
            if gender == 'M':
                flight_info['totals']['male'] += 1
            else:
                flight_info['totals']['female'] += 1
    
    # Parse totals from the file
    for i, line in enumerate(lines):
        if 'TOTALS PASSENGERS:' in line or (line.strip().startswith('.') and 'TOTALS:' in lines[i-2] if i >= 2 else False):
            # Look for the totals line
            totals_match = re.search(r'\.\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)', line)
            if totals_match:
                flight_info['totals']['male'] = int(totals_match.group(1))
                flight_info['totals']['female'] = int(totals_match.group(2))
                flight_info['totals']['child'] = int(totals_match.group(3))
                flight_info['totals']['infant'] = int(totals_match.group(4))
                flight_info['totals']['bags'] = int(totals_match.group(5))
                flight_info['totals']['weight'] = int(totals_match.group(6))
    
    flight_info['totals']['total'] = len(flight_info['passengers'])
    
    return flight_info

@manifest_bp.route('/manifest-dashboard')
def manifest_dashboard():
    """Manifest upload and analytics dashboard"""
    return render_template('manifest-dashboard.html')

@manifest_bp.route('/forecast-interface')
def forecast_interface():
    """Manual forecast data entry interface"""
    return render_template('forecast-interface.html')

@manifest_bp.route('/manifest/upload', methods=['POST'])
def upload_manifest():
    """
    Upload daily manifest (actual passenger data)
    Supports both text (.txt) and Excel (.xlsx) formats
    """
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    try:
        filename = file.filename.lower()
        
        # Handle text manifest files
        if filename.endswith('.txt'):
            file_content = file.read().decode('utf-8', errors='ignore')
            manifest_data = parse_text_manifest(file_content)
            
            if not manifest_data['flight_number'] or not manifest_data['date']:
                return jsonify({
                    'success': False, 
                    'error': 'Could not parse flight number or date from manifest'
                }), 400
            
            # Parse the date
            flight_date = datetime.strptime(manifest_data['date'], '%Y-%m-%d').date()
            flight_no = manifest_data['flight_number']
            
            # Get totals
            total_pax = manifest_data['totals']['total']
            male_count = manifest_data['totals']['male']
            female_count = manifest_data['totals']['female']
            
            # Get capacity (standard aircraft config for ET621/ET620)
            total_cap = 270  # Boeing 787 typical config
            business_cap = 24
            economy_cap = 246
            
            # Calculate load factors
            lf = (total_pax / total_cap * 100) if total_cap > 0 else 0
            
            # Convert route breakdown to regular dict
            route_breakdown = dict(manifest_data['route_breakdown'])
            
            # Determine direction based on flight number
            direction = 'outbound' if '621' in flight_no else 'inbound'
            
            # Check if manifest already exists for this flight/date
            existing = DailyManifest.query.filter_by(
                flight_date=flight_date,
                flight_number=flight_no
            ).first()
            
            if existing:
                # Update existing manifest
                existing.total_passengers = total_pax
                existing.business_passengers = 0  # Not parsed from text manifest
                existing.economy_passengers = total_pax
                existing.total_capacity = total_cap
                existing.business_capacity = business_cap
                existing.economy_capacity = economy_cap
                existing.load_factor = lf
                existing.business_load_factor = 0
                existing.economy_load_factor = (total_pax / economy_cap * 100) if economy_cap > 0 else 0
                existing.route_breakdown = route_breakdown
                existing.uploaded_at = datetime.utcnow()
                existing.uploaded_by = session.get('admin_username', 'admin')
                existing.source = 'manifest'
            else:
                # Create new manifest record
                new_manifest = DailyManifest(
                    flight_date=flight_date,
                    flight_number=flight_no,
                    direction=direction,
                    total_passengers=total_pax,
                    business_passengers=0,
                    economy_passengers=total_pax,
                    total_capacity=total_cap,
                    business_capacity=business_cap,
                    economy_capacity=economy_cap,
                    load_factor=lf,
                    business_load_factor=0,
                    economy_load_factor=(total_pax / economy_cap * 100) if economy_cap > 0 else 0,
                    route_breakdown=route_breakdown,
                    uploaded_by=session.get('admin_username', 'admin'),
                    source='manifest'
                )
                db.session.add(new_manifest)
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f'Successfully processed manifest for {flight_no} on {manifest_data["date"]}',
                'records_processed': 1,
                'flight_number': flight_no,
                'flight_date': manifest_data['date'],
                'total_passengers': total_pax,
                'route_breakdown': route_breakdown,
                'load_factor': round(lf, 1)
            })
        
        # Handle Excel files
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            file_content = file.read()
            workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
            sheet = workbook.active
            
            # Parse manifest data
            records_processed = 0
            
            for row_idx in range(2, sheet.max_row + 1):
                date_val = sheet.cell(row_idx, 1).value
                flight_no = str(sheet.cell(row_idx, 2).value) if sheet.cell(row_idx, 2).value else None
                direction = str(sheet.cell(row_idx, 3).value).lower() if sheet.cell(row_idx, 3).value else None
                total_pax = int(sheet.cell(row_idx, 4).value) if sheet.cell(row_idx, 4).value else 0
                business_pax = int(sheet.cell(row_idx, 5).value) if sheet.cell(row_idx, 5).value else 0
                economy_pax = int(sheet.cell(row_idx, 6).value) if sheet.cell(row_idx, 6).value else 0
                
                if not date_val or not flight_no:
                    continue
                
                # Convert date
                if isinstance(date_val, datetime):
                    flight_date = date_val.date()
                else:
                    flight_date = datetime.strptime(str(date_val), '%Y-%m-%d').date()
                
                # Get capacity
                total_cap = 270
                business_cap = 24
                economy_cap = 246
                
                # Calculate load factors
                lf = (total_pax / total_cap * 100) if total_cap > 0 else 0
                lf_c = (business_pax / business_cap * 100) if business_cap > 0 else 0
                lf_y = (economy_pax / economy_cap * 100) if economy_cap > 0 else 0
                
                route_breakdown = {}
                
                # Check if manifest already exists
                existing = DailyManifest.query.filter_by(
                    flight_date=flight_date,
                    flight_number=flight_no
                ).first()
                
                if existing:
                    existing.total_passengers = total_pax
                    existing.business_passengers = business_pax
                    existing.economy_passengers = economy_pax
                    existing.total_capacity = total_cap
                    existing.business_capacity = business_cap
                    existing.economy_capacity = economy_cap
                    existing.load_factor = lf
                    existing.business_load_factor = lf_c
                    existing.economy_load_factor = lf_y
                    existing.route_breakdown = route_breakdown
                    existing.uploaded_at = datetime.utcnow()
                    existing.uploaded_by = session.get('admin_username', 'admin')
                    existing.source = 'manifest'
                else:
                    new_manifest = DailyManifest(
                        flight_date=flight_date,
                        flight_number=flight_no,
                        direction=direction or ('inbound' if '620' in flight_no else 'outbound'),
                        total_passengers=total_pax,
                        business_passengers=business_pax,
                        economy_passengers=economy_pax,
                        total_capacity=total_cap,
                        business_capacity=business_cap,
                        economy_capacity=economy_cap,
                        load_factor=lf,
                        business_load_factor=lf_c,
                        economy_load_factor=lf_y,
                        route_breakdown=route_breakdown,
                        uploaded_by=session.get('admin_username', 'admin'),
                        source='manifest'
                    )
                    db.session.add(new_manifest)
                
                records_processed += 1
            
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f'Successfully processed {records_processed} manifest records',
                'records_processed': records_processed
            })
        
        else:
            return jsonify({
                'success': False, 
                'error': 'Unsupported file format. Please upload .txt or .xlsx files'
            }), 400
    
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@manifest_bp.route('/manifest/data')
def get_manifest_data():
    """Get manifest data for date range"""
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    flight_number = request.args.get('flight_number')
    
    query = DailyManifest.query
    
    if start_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        query = query.filter(DailyManifest.flight_date >= start_date)
    
    if end_date_str:
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        query = query.filter(DailyManifest.flight_date <= end_date)
    
    if flight_number:
        query = query.filter(DailyManifest.flight_number.like(f'%{flight_number}%'))
    
    records = query.order_by(DailyManifest.flight_date).all()
    
    return jsonify({
        'success': True,
        'data': [r.to_dict() for r in records],
        'total_passengers': sum(r.total_passengers for r in records),
        'average_load_factor': sum(r.load_factor for r in records) / len(records) if records else 0,
        'record_count': len(records)
    })

@manifest_bp.route('/forecast/save', methods=['POST'])
def save_forecast():
    """
    Save manual forecast data
    """
    data = request.get_json()
    forecasts = data.get('forecasts', [])
    
    try:
        saved_count = 0
        
        for forecast in forecasts:
            forecast_date = datetime.strptime(forecast['date'], '%Y-%m-%d').date()
            airport_code = forecast['airport_code']
            direction = forecast.get('direction', 'outbound')
            passengers = int(forecast['passengers'])
            
            # Check if forecast exists
            existing = RouteForecast.query.filter_by(
                forecast_date=forecast_date,
                airport_code=airport_code,
                direction=direction
            ).first()
            
            if existing:
                existing.passengers = passengers
                existing.updated_at = datetime.utcnow()
                existing.created_by = session.get('admin_username', 'admin')
            else:
                new_forecast = RouteForecast(
                    forecast_date=forecast_date,
                    airport_code=airport_code,
                    direction=direction,
                    passengers=passengers,
                    created_by=session.get('admin_username', 'admin')
                )
                db.session.add(new_forecast)
            
            saved_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully saved {saved_count} forecasts'
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@manifest_bp.route('/forecast/data')
def get_forecast_data():
    """
    Get combined forecast and manifest data
    """
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    direction = request.args.get('direction', 'outbound')
    
    if not start_date_str or not end_date_str:
        return jsonify({'success': False, 'error': 'Date range required'}), 400
    
    start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    
    # Get all dates in range
    date_list = []
    current_date = start_date
    while current_date <= end_date:
        date_list.append(current_date)
        current_date += timedelta(days=1)
    
    # Get forecasts for date range
    forecasts = RouteForecast.query.filter(
        RouteForecast.forecast_date >= start_date,
        RouteForecast.forecast_date <= end_date,
        RouteForecast.direction == direction
    ).all()
    
    # Get manifests for date range (actuals)
    manifests = DailyManifest.query.filter(
        DailyManifest.flight_date >= start_date,
        DailyManifest.flight_date <= end_date,
        DailyManifest.direction == direction
    ).all()
    
    # Build combined data structure
    data_by_airport = defaultdict(lambda: defaultdict(lambda: {'passengers': 0, 'source': 'forecast', 'confirmed': False}))
    
    # First, add forecast data
    for forecast in forecasts:
        date_str = forecast.forecast_date.strftime('%Y-%m-%d')
        data_by_airport[forecast.airport_code][date_str] = {
            'passengers': forecast.passengers,
            'source': 'forecast',
            'confirmed': False
        }
    
    # Then, override with manifest data (actuals) where available
    manifest_dates = set()
    for manifest in manifests:
        date_str = manifest.flight_date.strftime('%Y-%m-%d')
        manifest_dates.add(date_str)
        
        # Parse route breakdown
        if manifest.route_breakdown:
            for airport, pax_count in manifest.route_breakdown.items():
                data_by_airport[airport][date_str] = {
                    'passengers': pax_count,
                    'source': 'manifest',
                    'confirmed': True
                }
    
    # Get all airports
    airports = AirportMaster.query.filter_by(active=True).all()
    airport_codes = [a.code for a in airports]
    
    # Add any airports from data that aren't in master list
    for airport in data_by_airport.keys():
        if airport not in airport_codes:
            airport_codes.append(airport)
    
    # Build response
    result = {
        'dates': [d.strftime('%Y-%m-%d') for d in date_list],
        'airports': sorted(airport_codes),
        'data': dict(data_by_airport),
        'manifest_dates': list(manifest_dates)
    }
    
    return jsonify({
        'success': True,
        'result': result
    })

@manifest_bp.route('/airports/list')
def list_airports():
    """Get list of airports for dropdown"""
    airports = AirportMaster.query.filter_by(active=True).order_by(AirportMaster.code).all()
    return jsonify({
        'success': True,
        'airports': [a.to_dict() for a in airports]
    })

@manifest_bp.route('/airports/add', methods=['POST'])
def add_airport():
    """Add new airport to master list"""
    data = request.get_json()
    code = data.get('code', '').upper().strip()
    name = data.get('name', '').strip()
    country = data.get('country', '').strip()
    
    if not code:
        return jsonify({'success': False, 'error': 'Airport code required'}), 400
    
    # Check if exists
    existing = AirportMaster.query.filter_by(code=code).first()
    if existing:
        return jsonify({'success': False, 'error': 'Airport code already exists'}), 400
    
    try:
        new_airport = AirportMaster(
            code=code,
            name=name,
            country=country,
            active=True
        )
        db.session.add(new_airport)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'airport': new_airport.to_dict()
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
