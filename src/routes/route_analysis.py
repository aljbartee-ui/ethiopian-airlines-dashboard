"""
Complete Route Analysis Routes with Database Storage
Persistent data across server restarts
"""
from flask import Blueprint, request, jsonify
from werkzeug.utils import secure_filename
import openpyxl
import os
import time
from datetime import datetime
from src.models.user import db
from src.models.route_analysis import RouteAnalysisWeek, RouteAnalysisUpload
from src.utils.airport_lookup import get_airport_info
import json

route_analysis_bp = Blueprint('route_analysis', __name__)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def parse_route_section(ws, header_row, start_col, end_col, section_type):
    """
    Parse either outbound or inbound section
    section_type: 'outbound' or 'inbound'
    """
    # Extract dates from header
    dates = []
    col_idx = start_col + 1  # Skip POINTS column
    while col_idx <= end_col:
        cell_value = ws.cell(header_row, col_idx).value
        if cell_value is None:
            break
        # Stop at GRAND TOTAL column
        if isinstance(cell_value, str) and 'TOTAL' in cell_value.upper():
            break
        if isinstance(cell_value, datetime):
            dates.append(cell_value.strftime('%Y-%m-%d'))
        else:
            try:
                date_str = str(cell_value).split()[0]
                dates.append(date_str)
            except:
                pass
        col_idx += 1
    
    # Extract routes
    routes = []
    row_idx = header_row + 1
    
    while row_idx <= ws.max_row:
        route_code = ws.cell(row_idx, start_col).value
        
        if not route_code:
            row_idx += 1
            continue
        
        route_code = str(route_code).strip().upper()
        
        # Skip TOTAL row
        if 'TOTAL' in route_code:
            row_idx += 1
            continue
        
        # Extract daily passengers
        daily_passengers = {}
        total_passengers = 0
        
        for i, date in enumerate(dates):
            cell_value = ws.cell(row_idx, start_col + 1 + i).value
            passengers = 0
            if cell_value is not None:
                try:
                    passengers = int(float(cell_value))
                except:
                    passengers = 0
            daily_passengers[date] = passengers
            total_passengers += passengers
        
        # Only include routes with passengers
        if total_passengers > 0:
            # Get airport info or use code as-is
            airport_info = get_airport_info(route_code)
            if airport_info:
                display_name = f"{route_code} - {airport_info['city']}, {airport_info['country']}"
            else:
                # Use airport code as-is if not identified
                display_name = route_code
            
            routes.append({
                'code': route_code,
                'display_name': display_name,
                'total_passengers': total_passengers,
                'daily_passengers': daily_passengers,
                'type': section_type
            })
        
        row_idx += 1
    
    return routes, dates

def parse_route_analysis_excel_all_sheets(file_path):
    """
    Parse ALL sheets with outbound/inbound separation
    Outbound: Columns A-J (destinations from ADD)
    Inbound: Columns K-T (origins to ADD)
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        all_weeks_data = []
        skipped_count = 0
        
        print(f"\nProcessing {len(wb.sheetnames)} sheets...")
        
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                
                # Find header row (should be row 2 with "POINTS")
                header_row = None
                for row_idx in range(1, 5):
                    first_cell = ws.cell(row_idx, 1).value
                    if first_cell and 'POINTS' in str(first_cell).upper():
                        header_row = row_idx
                        break
                
                if not header_row:
                    print(f"  ⚠ Skipping {sheet_name}: No POINTS header")
                    skipped_count += 1
                    continue
                
                # Parse OUTBOUND section (columns A-J, or 1-10)
                outbound_routes, outbound_dates = parse_route_section(
                    ws, header_row, start_col=1, end_col=10, section_type='outbound'
                )
                
                # Parse INBOUND section (columns K-T, or 11-20)
                inbound_routes, inbound_dates = parse_route_section(
                    ws, header_row, start_col=11, end_col=20, section_type='inbound'
                )
                
                if not outbound_routes and not inbound_routes:
                    print(f"  ⚠ Skipping {sheet_name}: No route data")
                    skipped_count += 1
                    continue
                
                # Calculate totals
                outbound_total = sum(r['total_passengers'] for r in outbound_routes)
                inbound_total = sum(r['total_passengers'] for r in inbound_routes)
                total_passengers = outbound_total + inbound_total
                
                # Find top routes
                top_destination = max(outbound_routes, key=lambda x: x['total_passengers']) if outbound_routes else None
                top_origin = max(inbound_routes, key=lambda x: x['total_passengers']) if inbound_routes else None
                
                # Find busiest days
                outbound_daily_totals = {}
                for route in outbound_routes:
                    for date, pax in route['daily_passengers'].items():
                        outbound_daily_totals[date] = outbound_daily_totals.get(date, 0) + pax
                
                inbound_daily_totals = {}
                for route in inbound_routes:
                    for date, pax in route['daily_passengers'].items():
                        inbound_daily_totals[date] = inbound_daily_totals.get(date, 0) + pax
                
                outbound_busiest = max(outbound_daily_totals.items(), key=lambda x: x[1]) if outbound_daily_totals else (None, 0)
                inbound_busiest = max(inbound_daily_totals.items(), key=lambda x: x[1]) if inbound_daily_totals else (None, 0)
                
                week_data = {
                    'sheet_name': sheet_name,
                    'outbound': {
                        'routes': outbound_routes,
                        'total_routes': len(outbound_routes),
                        'total_passengers': outbound_total,
                        'dates': outbound_dates,
                        'top_route': top_destination,
                        'busiest_day': {
                            'date': outbound_busiest[0],
                            'passengers': outbound_busiest[1]
                        } if outbound_busiest[0] else None
                    },
                    'inbound': {
                        'routes': inbound_routes,
                        'total_routes': len(inbound_routes),
                        'total_passengers': inbound_total,
                        'dates': inbound_dates,
                        'top_route': top_origin,
                        'busiest_day': {
                            'date': inbound_busiest[0],
                            'passengers': inbound_busiest[1]
                        } if inbound_busiest[0] else None
                    },
                    'total_passengers': total_passengers,
                    'total_routes': len(outbound_routes) + len(inbound_routes)
                }
                
                all_weeks_data.append(week_data)
                print(f"  ✓ {sheet_name}: OUT={len(outbound_routes)} routes ({outbound_total} pax), IN={len(inbound_routes)} routes ({inbound_total} pax)")
            
            except Exception as e:
                print(f"  ✗ Error in {sheet_name}: {str(e)}")
                skipped_count += 1
                continue
        
        wb.close()
        
        print(f"\n✅ Successfully processed {len(all_weeks_data)} weeks (skipped {skipped_count})\n")
        
        return {
            'success': True,
            'weeks': all_weeks_data,
            'total_weeks': len(all_weeks_data),
            'skipped_weeks': skipped_count
        }
    
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

@route_analysis_bp.route('/upload', methods=['POST'])
def upload_route_analysis():
    """Upload and process route analysis Excel file - ALL SHEETS with database storage"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'success': False, 'error': 'Invalid file type. Please upload .xlsx or .xls file'}), 400
    
    try:
        start_time = time.time()
        
        # Save file temporarily
        filename = secure_filename(file.filename)
        filepath = os.path.join('/tmp', filename)
        file.save(filepath)
        
        # Parse ALL sheets
        result = parse_route_analysis_excel_all_sheets(filepath)
        
        if not result['success']:
            return jsonify({'success': False, 'error': result['error']}), 400
        
        # Clear existing data
        RouteAnalysisWeek.query.delete()
        db.session.commit()
        
        # Store each week's data in database
        for week_data in result['weeks']:
            # Create database record
            week_record = RouteAnalysisWeek(
                sheet_name=week_data['sheet_name'],
                
                # Outbound data
                outbound_routes=json.dumps(week_data['outbound']['routes']),
                outbound_total_routes=week_data['outbound']['total_routes'],
                outbound_total_passengers=week_data['outbound']['total_passengers'],
                outbound_dates=json.dumps(week_data['outbound']['dates']),
                outbound_top_route_code=week_data['outbound']['top_route']['code'] if week_data['outbound']['top_route'] else None,
                outbound_top_route_name=week_data['outbound']['top_route']['display_name'] if week_data['outbound']['top_route'] else None,
                outbound_top_route_passengers=week_data['outbound']['top_route']['total_passengers'] if week_data['outbound']['top_route'] else 0,
                outbound_busiest_day=week_data['outbound']['busiest_day']['date'] if week_data['outbound']['busiest_day'] else None,
                outbound_busiest_day_passengers=week_data['outbound']['busiest_day']['passengers'] if week_data['outbound']['busiest_day'] else 0,
                
                # Inbound data
                inbound_routes=json.dumps(week_data['inbound']['routes']),
                inbound_total_routes=week_data['inbound']['total_routes'],
                inbound_total_passengers=week_data['inbound']['total_passengers'],
                inbound_dates=json.dumps(week_data['inbound']['dates']),
                inbound_top_route_code=week_data['inbound']['top_route']['code'] if week_data['inbound']['top_route'] else None,
                inbound_top_route_name=week_data['inbound']['top_route']['display_name'] if week_data['inbound']['top_route'] else None,
                inbound_top_route_passengers=week_data['inbound']['top_route']['total_passengers'] if week_data['inbound']['top_route'] else 0,
                inbound_busiest_day=week_data['inbound']['busiest_day']['date'] if week_data['inbound']['busiest_day'] else None,
                inbound_busiest_day_passengers=week_data['inbound']['busiest_day']['passengers'] if week_data['inbound']['busiest_day'] else 0,
                
                # Combined totals
                total_passengers=week_data['total_passengers'],
                total_routes=week_data['total_routes'],
                upload_date=datetime.now()
            )
            
            db.session.add(week_record)
        
        db.session.commit()
        
        # Record upload history
        processing_time = time.time() - start_time
        upload_record = RouteAnalysisUpload(
            filename=filename,
            total_weeks_processed=result['total_weeks'],
            total_weeks_skipped=result['skipped_weeks'],
            processing_time_seconds=processing_time
        )
        db.session.add(upload_record)
        db.session.commit()
        
        # Clean up
        os.remove(filepath)
        
        return jsonify({
            'success': True,
            'message': f'Successfully processed {result["total_weeks"]} weeks of data',
            'total_weeks': result['total_weeks'],
            'skipped_weeks': result['skipped_weeks'],
            'processing_time': round(processing_time, 2)
        })
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@route_analysis_bp.route('/weeks', methods=['GET'])
def get_available_weeks():
    """Get list of all available weeks from database"""
    try:
        weeks = RouteAnalysisWeek.query.order_by(RouteAnalysisWeek.id.desc()).all()
        
        week_list = [{
            'id': week.id,
            'name': week.sheet_name,
            'total_routes': week.total_routes,
            'total_passengers': week.total_passengers,
            'outbound_passengers': week.outbound_total_passengers,
            'inbound_passengers': week.inbound_total_passengers
        } for week in weeks]
        
        return jsonify({
            'success': True,
            'weeks': week_list
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@route_analysis_bp.route('/data', methods=['GET'])
def get_route_analysis_data():
    """Get route analysis data for a specific week from database"""
    week_id = request.args.get('week_id', type=int)
    
    try:
        if week_id:
            week = RouteAnalysisWeek.query.get(week_id)
        else:
            # Get most recent week
            week = RouteAnalysisWeek.query.order_by(RouteAnalysisWeek.id.desc()).first()
        
        if not week:
            return jsonify({'success': False, 'error': 'No data available'})
        
        return jsonify({
            'success': True,
            'week_id': week.id,
            'summary': {
                'week': week.sheet_name,
                'total_passengers': week.total_passengers,
                'total_routes': week.total_routes,
                'outbound': {
                    'total_routes': week.outbound_total_routes,
                    'total_passengers': week.outbound_total_passengers,
                    'top_destination': {
                        'code': week.outbound_top_route_code,
                        'display_name': week.outbound_top_route_name,
                        'passengers': week.outbound_top_route_passengers
                    } if week.outbound_top_route_code else None,
                    'busiest_day': {
                        'date': week.outbound_busiest_day,
                        'passengers': week.outbound_busiest_day_passengers
                    } if week.outbound_busiest_day else None
                },
                'inbound': {
                    'total_routes': week.inbound_total_routes,
                    'total_passengers': week.inbound_total_passengers,
                    'top_origin': {
                        'code': week.inbound_top_route_code,
                        'display_name': week.inbound_top_route_name,
                        'passengers': week.inbound_top_route_passengers
                    } if week.inbound_top_route_code else None,
                    'busiest_day': {
                        'date': week.inbound_busiest_day,
                        'passengers': week.inbound_busiest_day_passengers
                    } if week.inbound_busiest_day else None
                }
            }
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@route_analysis_bp.route('/compare', methods=['GET'])
def compare_weeks():
    """Compare two weeks from database"""
    week1_id = request.args.get('week1_id', type=int)
    week2_id = request.args.get('week2_id', type=int)
    
    if not week1_id or not week2_id:
        return jsonify({'success': False, 'error': 'Both week IDs required'}), 400
    
    try:
        week1 = RouteAnalysisWeek.query.get(week1_id)
        week2 = RouteAnalysisWeek.query.get(week2_id)
        
        if not week1 or not week2:
            return jsonify({'success': False, 'error': 'Week not found'}), 404
        
        variance = week1.total_passengers - week2.total_passengers
        variance_pct = round((variance / week2.total_passengers * 100), 2) if week2.total_passengers > 0 else 0
        
        return jsonify({
            'success': True,
            'comparison': {
                'week1': {
                    'id': week1.id,
                    'name': week1.sheet_name,
                    'passengers': week1.total_passengers
                },
                'week2': {
                    'id': week2.id,
                    'name': week2.sheet_name,
                    'passengers': week2.total_passengers
                },
                'variance': variance,
                'variance_pct': variance_pct
            }
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Chart endpoints continue in next part...
@route_analysis_bp.route('/charts/top-destinations', methods=['GET'])
def get_top_destinations_chart():
    """Get top destinations (outbound) chart from database"""
    week_id = request.args.get('week_id', type=int)
    limit = request.args.get('limit', 10, type=int)
    
    try:
        if week_id:
            week = RouteAnalysisWeek.query.get(week_id)
        else:
            week = RouteAnalysisWeek.query.order_by(RouteAnalysisWeek.id.desc()).first()
        
        if not week:
            return jsonify({'success': False, 'error': 'No data available'})
        
        routes = json.loads(week.outbound_routes)
        top_routes = sorted(routes, key=lambda x: x['total_passengers'], reverse=True)[:limit]
        
        labels = [r['display_name'] for r in top_routes]
        values = [r['total_passengers'] for r in top_routes]
        
        return jsonify({
            'success': True,
            'labels': labels,
            'datasets': [{'label': 'Passengers', 'data': values}]
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@route_analysis_bp.route('/charts/top-origins', methods=['GET'])
def get_top_origins_chart():
    """Get top origins (inbound) chart from database"""
    week_id = request.args.get('week_id', type=int)
    limit = request.args.get('limit', 10, type=int)
    
    try:
        if week_id:
            week = RouteAnalysisWeek.query.get(week_id)
        else:
            week = RouteAnalysisWeek.query.order_by(RouteAnalysisWeek.id.desc()).first()
        
        if not week:
            return jsonify({'success': False, 'error': 'No data available'})
        
        routes = json.loads(week.inbound_routes)
        top_routes = sorted(routes, key=lambda x: x['total_passengers'], reverse=True)[:limit]
        
        labels = [r['display_name'] for r in top_routes]
        values = [r['total_passengers'] for r in top_routes]
        
        return jsonify({
            'success': True,
            'labels': labels,
            'datasets': [{'label': 'Passengers', 'data': values}]
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@route_analysis_bp.route('/charts/outbound-vs-inbound', methods=['GET'])
def get_outbound_vs_inbound_chart():
    """Compare outbound vs inbound traffic from database"""
    week_id = request.args.get('week_id', type=int)
    
    try:
        if week_id:
            week = RouteAnalysisWeek.query.get(week_id)
        else:
            week = RouteAnalysisWeek.query.order_by(RouteAnalysisWeek.id.desc()).first()
        
        if not week:
            return jsonify({'success': False, 'error': 'No data available'})
        
        return jsonify({
            'success': True,
            'labels': ['Outbound (Destinations)', 'Inbound (Origins)'],
            'datasets': [{
                'label': 'Passengers',
                'data': [week.outbound_total_passengers, week.inbound_total_passengers]
            }]
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@route_analysis_bp.route('/charts/week-comparison', methods=['GET'])
def get_week_comparison_chart():
    """Compare two weeks from database"""
    week1_id = request.args.get('week1_id', type=int)
    week2_id = request.args.get('week2_id', type=int)
    limit = request.args.get('limit', 10, type=int)
    direction = request.args.get('direction', 'outbound')
    
    try:
        if not week1_id or not week2_id:
            # Default to latest two weeks
            weeks = RouteAnalysisWeek.query.order_by(RouteAnalysisWeek.id.desc()).limit(2).all()
            if len(weeks) < 2:
                return jsonify({'success': False, 'error': 'Need at least 2 weeks'})
            week1, week2 = weeks[0], weeks[1]
        else:
            week1 = RouteAnalysisWeek.query.get(week1_id)
            week2 = RouteAnalysisWeek.query.get(week2_id)
        
        if not week1 or not week2:
            return jsonify({'success': False, 'error': 'Week not found'})
        
        routes1 = json.loads(week1.outbound_routes if direction == 'outbound' else week1.inbound_routes)
        routes2 = json.loads(week2.outbound_routes if direction == 'outbound' else week2.inbound_routes)
        
        # Get top routes from week1
        top_routes = sorted(routes1, key=lambda x: x['total_passengers'], reverse=True)[:limit]
        
        # Create lookup for week2
        week2_lookup = {r['code']: r['total_passengers'] for r in routes2}
        
        labels = [r['display_name'] for r in top_routes]
        week1_values = [r['total_passengers'] for r in top_routes]
        week2_values = [week2_lookup.get(r['code'], 0) for r in top_routes]
        
        return jsonify({
            'success': True,
            'labels': labels,
            'datasets': [
                {'label': week1.sheet_name, 'data': week1_values},
                {'label': week2.sheet_name, 'data': week2_values}
            ]
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@route_analysis_bp.route('/charts/growth-rates', methods=['GET'])
def get_growth_rates_chart():
    """Get growth rates comparing two weeks from database"""
    week1_id = request.args.get('week1_id', type=int)
    week2_id = request.args.get('week2_id', type=int)
    limit = request.args.get('limit', 10, type=int)
    direction = request.args.get('direction', 'outbound')
    
    try:
        if not week1_id or not week2_id:
            # Default to latest two weeks
            weeks = RouteAnalysisWeek.query.order_by(RouteAnalysisWeek.id.desc()).limit(2).all()
            if len(weeks) < 2:
                return jsonify({'success': False, 'error': 'Need at least 2 weeks'})
            week1, week2 = weeks[0], weeks[1]
        else:
            week1 = RouteAnalysisWeek.query.get(week1_id)
            week2 = RouteAnalysisWeek.query.get(week2_id)
        
        if not week1 or not week2:
            return jsonify({'success': False, 'error': 'Week not found'})
        
        routes1 = json.loads(week1.outbound_routes if direction == 'outbound' else week1.inbound_routes)
        routes2 = json.loads(week2.outbound_routes if direction == 'outbound' else week2.inbound_routes)
        
        # Create lookup for week2
        week2_lookup = {r['code']: r['total_passengers'] for r in routes2}
        
        # Calculate growth rates
        growth_data = []
        for route in routes1:
            code = route['code']
            current = route['total_passengers']
            previous = week2_lookup.get(code, 0)
            
            if previous > 0:
                growth_rate = ((current - previous) / previous) * 100
                growth_data.append({
                    'display_name': route['display_name'],
                    'growth_rate': round(growth_rate, 2),
                    'current': current,
                    'previous': previous
                })
        
        # Sort by absolute growth and get top N
        growth_data.sort(key=lambda x: abs(x['growth_rate']), reverse=True)
        top_growth = growth_data[:limit]
        
        labels = [r['display_name'] for r in top_growth]
        values = [r['growth_rate'] for r in top_growth]
        
        return jsonify({
            'success': True,
            'labels': labels,
            'datasets': [{'label': 'Growth Rate (%)', 'data': values}]
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@route_analysis_bp.route('/charts/daily-trend', methods=['GET'])
def get_daily_trend_chart():
    """Get daily passenger trend from database"""
    week_id = request.args.get('week_id', type=int)
    direction = request.args.get('direction', 'both')
    
    try:
        if week_id:
            week = RouteAnalysisWeek.query.get(week_id)
        else:
            week = RouteAnalysisWeek.query.order_by(RouteAnalysisWeek.id.desc()).first()
        
        if not week:
            return jsonify({'success': False, 'error': 'No data available'})
        
        datasets = []
        
        if direction in ['outbound', 'both']:
            routes = json.loads(week.outbound_routes)
            dates = json.loads(week.outbound_dates)
            daily_totals = {date: 0 for date in dates}
            for route in routes:
                for date, pax in route['daily_passengers'].items():
                    if date in daily_totals:
                        daily_totals[date] += pax
            
            formatted_dates = [datetime.strptime(d, '%Y-%m-%d').strftime('%b %d') for d in dates]
            values = [daily_totals[d] for d in dates]
            
            datasets.append({
                'label': 'Outbound',
                'data': values
            })
        
        if direction in ['inbound', 'both']:
            routes = json.loads(week.inbound_routes)
            dates = json.loads(week.inbound_dates)
            daily_totals = {date: 0 for date in dates}
            for route in routes:
                for date, pax in route['daily_passengers'].items():
                    if date in daily_totals:
                        daily_totals[date] += pax
            
            formatted_dates = [datetime.strptime(d, '%Y-%m-%d').strftime('%b %d') for d in dates]
            values = [daily_totals[d] for d in dates]
            
            datasets.append({
                'label': 'Inbound',
                'data': values
            })
        
        return jsonify({
            'success': True,
            'labels': formatted_dates,
            'datasets': datasets
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
