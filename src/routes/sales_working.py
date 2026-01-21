from flask import Blueprint, request, jsonify, session
from src.models.user import db
from src.models.sales import SalesData, AdminUser
import os
import json
from datetime import datetime
import base64
import openpyxl
from io import BytesIO

sales_bp = Blueprint('sales', __name__)

def process_excel_file(file_content, filename):
    """Process Excel file and extract data"""
    try:
        # Load workbook from bytes
        workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        
        processed_data = {}
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Get headers from first row and clean them (strip whitespace)
            headers = []
            for cell in sheet[1]:
                header = cell.value if cell.value is not None else ''
                # Clean header - strip whitespace
                if isinstance(header, str):
                    header = header.strip()
                headers.append(header)
            
            # Skip sheets with no meaningful headers
            if not any(headers):
                continue
            
            # Get data rows
            data_rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(headers) and headers[i]:
                            # Convert datetime objects to strings
                            if hasattr(value, 'strftime'):
                                value = value.strftime('%Y-%m-%d %H:%M:%S')
                            # Clean the header name for the key
                            clean_header = headers[i].strip() if isinstance(headers[i], str) else headers[i]
                            row_dict[clean_header] = value
                    if row_dict:  # Only add non-empty rows
                        data_rows.append(row_dict)
            
            # Only add sheets that have data
            if data_rows:
                processed_data[sheet_name] = {
                    'headers': headers,
                    'data': data_rows,
                    'row_count': len(data_rows)
                }
        
        # If we have a sheet with actual data, prioritize it
        # Look for sheets with meaningful data (more than 10 rows typically)
        best_sheet = None
        max_rows = 0
        for sheet_name, sheet_data in processed_data.items():
            if sheet_data['row_count'] > max_rows:
                max_rows = sheet_data['row_count']
                best_sheet = sheet_name
        
        # If we found a best sheet with significant data, use it as primary
        if best_sheet and max_rows > 10:
            # Reorder to put best sheet first
            reordered = {best_sheet: processed_data[best_sheet]}
            for k, v in processed_data.items():
                if k != best_sheet:
                    reordered[k] = v
            processed_data = reordered
        
        return processed_data
        
    except Exception as e:
        print(f"Error processing Excel file: {e}")
        raise e

@sales_bp.route('/data')
def get_current_data():
    """Get information about the current active dataset"""
    try:
        active_data = SalesData.query.filter_by(is_active=True).first()
        if active_data:
            data = active_data.get_data()
            sheets = list(data.keys()) if data else []
            return jsonify({
                'filename': active_data.filename,
                'upload_date': active_data.upload_date.isoformat(),
                'sheets': sheets,
                'total_rows': sum(sheet_data.get('row_count', 0) for sheet_data in data.values()) if data else 0
            })
        else:
            return jsonify({'error': 'No data available'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@sales_bp.route('/upload', methods=['POST'])
def upload_file():
    """Handle Excel file upload (admin only)"""
    if not session.get('admin_logged_in'):
        return jsonify({'error': 'Admin authentication required'}), 401
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Invalid file type. Please upload Excel files only.'}), 400
    
    try:
        # Read file content
        file_content = file.read()
        
        # Process Excel file
        processed_data = process_excel_file(file_content, file.filename)
        
        if not processed_data:
            return jsonify({'error': 'No data found in Excel file'}), 400
        
        # Deactivate all previous data
        SalesData.query.update({'is_active': False})
        
        # Create new sales data entry
        sales_data = SalesData(
            filename=file.filename,
            data_json=json.dumps(processed_data),
            is_active=True
        )
        
        db.session.add(sales_data)
        db.session.commit()
        
        # Calculate summary statistics
        total_rows = sum(sheet_data.get('row_count', 0) for sheet_data in processed_data.values())
        sheets = list(processed_data.keys())
        
        return jsonify({
            'message': 'File uploaded and processed successfully',
            'filename': file.filename,
            'data_id': sales_data.id,
            'sheets': sheets,
            'total_rows': total_rows,
            'summary': {
                'sheets_processed': len(sheets),
                'total_data_rows': total_rows
            }
        })
        
    except Exception as e:
        db.session.rollback()
        print(f"Upload error: {e}")
        return jsonify({'error': f'Error processing file: {str(e)}'}), 500

@sales_bp.route('/charts/default')
def generate_default_charts():
    """Generate default charts from the active dataset"""
    try:
        active_data = SalesData.query.filter_by(is_active=True).first()
        if not active_data:
            return jsonify({'error': 'No active dataset found'}), 404
        
        # Get the actual data
        data = active_data.get_data()
        if not data:
            return jsonify({'error': 'No data available'}), 404
        
        # Return success message for now - charts will be generated by the charts endpoint
        return jsonify({
            'message': 'Data is ready for chart generation',
            'sheets': list(data.keys()),
            'total_rows': sum(sheet_data.get('row_count', 0) for sheet_data in data.values())
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@sales_bp.route('/debug/data')
def debug_data():
    """Debug endpoint to check data structure"""
    try:
        active_data = SalesData.query.filter_by(is_active=True).first()
        if not active_data:
            return jsonify({'error': 'No active dataset found'}), 404
        
        data = active_data.get_data()
        
        # Return structure info
        debug_info = {}
        for sheet_name, sheet_data in data.items():
            debug_info[sheet_name] = {
                'headers': sheet_data.get('headers', []),
                'row_count': sheet_data.get('row_count', 0),
                'sample_row': sheet_data.get('data', [{}])[0] if sheet_data.get('data') else {}
            }
        
        return jsonify({
            'filename': active_data.filename,
            'upload_date': active_data.upload_date.isoformat(),
            'debug_info': debug_info
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
