from flask import Blueprint, request, jsonify, session
from src.models.user import db
from src.models.flight_load import FlightLoadRecord
import openpyxl
from io import BytesIO
from datetime import datetime
from collections import defaultdict

flight_load_bp = Blueprint('flight_load', __name__)

def safe_int(value):
    """Safely convert value to int, handling non-numeric values"""
    if value is None or value == '':
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    if isinstance(value, str):
        # Handle special values like 'X', 'N/A', etc.
        if value.strip().upper() in ['X', 'N/A', 'NA', '-', '']:
            return 0
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return 0
    return 0

def safe_float(value):
    """Safely convert value to float, handling non-numeric values"""
    if value is None or value == '':
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Handle special values like 'X', 'N/A', etc.
        if value.strip().upper() in ['X', 'N/A', 'NA', '-', '']:
            return 0.0
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0
    return 0.0

def process_flight_load_excel(file_content, filename):
    """Process Flight Load Excel file and extract LF 620-621 data"""
    try:
        # Load workbook with data_only=True to read formula values
        workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        
        # Get the first sheet (should be LF 620-621)
        if not workbook.sheetnames:
            raise ValueError("No sheets found in Excel file")
        
        # Use first sheet
        sheet = workbook[workbook.sheetnames[0]]
        print(f"Processing sheet: {workbook.sheetnames[0]}")
        
        processed_data = {
            'inbound': [],  # Flight 620: ADD to KWI
            'outbound': []  # Flight 621: KWI to ADD
        }
        
        # Process inbound flights (columns A to L)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If flight number exists
                inbound_record = {
                    'flight_no': str(row[0]),
                    'travel_date': row[1].strftime('%Y-%m-%d') if isinstance(row[1], datetime) else str(row[1]),
                    'day': str(row[2]) if row[2] else '',
                    'c_cap': safe_int(row[3]),  # Business capacity
                    'y_cap': safe_int(row[4]),  # Economy capacity
                    'tot_cap': safe_int(row[5]),  # Total capacity
                    'pax_c': safe_int(row[6]),  # Business passengers
                    'pax_y': safe_int(row[7]),  # Economy passengers
                    'pax': safe_int(row[8]),  # Total passengers
                    'lf_c': safe_float(row[9]),  # Business load factor
                    'lf_y': safe_float(row[10]),  # Economy load factor
                    'lf': safe_float(row[11])  # Total load factor
                }
                processed_data['inbound'].append(inbound_record)
        
        # Process outbound flights (columns O to AA, indices 14 to 26)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) > 14 and row[14]:  # If outbound flight number exists
                outbound_record = {
                    'flight_no': str(row[14]),
                    'travel_date': row[15].strftime('%Y-%m-%d') if isinstance(row[15], datetime) else str(row[15]),
                    'day': str(row[16]) if row[16] else '',
                    'c_cap': safe_int(row[17]),
                    'y_cap': safe_int(row[18]),
                    'tot_cap': safe_int(row[19]),
                    'pax_c': safe_int(row[20]),
                    'pax_y': safe_int(row[21]),
                    'pax': safe_int(row[22]),
                    'lf_c': safe_float(row[23]),
                    'lf_y': safe_float(row[24]),
                    'lf': safe_float(row[25])
                }
                processed_data['outbound'].append(outbound_record)
        
        return processed_data
        
    except Exception as e:
        print(f"Error processing flight load Excel file: {e}")
        raise e

@flight_load_bp.route('/upload', methods=['POST'])
def upload_flight_load():
    """Handle Load Factor Excel file upload (admin only) - Forecast Data"""
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Invalid file type. Please upload Excel files only.'}), 400
    
    try:
        # Read file content
        file_content = file.read()
        
        # Process Excel file
        processed_data = process_flight_load_excel(file_content, file.filename)
        
        if not processed_data['inbound'] and not processed_data['outbound']:
            return jsonify({'error': 'No flight load data found in Excel file'}), 400
        
        all_records = processed_data['inbound'] + processed_data['outbound']
        
        # Transaction to ensure atomicity
        with db.session.begin_nested():
            for record in all_records:
                travel_date = datetime.strptime(record['travel_date'], '%Y-%m-%d').date()
                flight_no = record['flight_no']
                
                # Check if an actual manifest record exists for this date/flight
                existing_record = FlightLoadRecord.query.filter_by(
                    travel_date=travel_date,
                    flight_no=flight_no
                ).first()
                
                if existing_record:
                    # If an actual manifest exists, DO NOT override it with forecast data
                    if existing_record.data_source == 'manifest':
                        continue
                    
                    # If it's an existing forecast, update it
                    existing_record.update_from_dict(record)
                    existing_record.data_source = 'forecast'
                else:
                    # Create a new forecast record
                    new_record = FlightLoadRecord(
                        travel_date=travel_date,
                        flight_no=flight_no,
                        data_source='forecast'
                    )
                    new_record.update_from_dict(record)
                    db.session.add(new_record)
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Load Factor (Forecast) data uploaded and merged successfully',
            'total_records_processed': len(all_records)
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Failed to process file: {str(e)}'}), 500

@flight_load_bp.route('/data')
def get_flight_load_data():
    """Get flight load data with optional filters"""
    # Check authentication
    is_admin = session.get('admin_logged_in', False)
    is_public = session.get('public_authenticated', False)
    
    if not is_admin and not is_public:
        return jsonify({'error': 'Authentication required'}), 401
    
    try:
        # Get filter parameters
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        flight_type = request.args.get('flight_type', 'both')  # inbound, outbound, or both
        
        query = FlightLoadRecord.query
        
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            query = query.filter(FlightLoadRecord.travel_date >= start_date)
        
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            query = query.filter(FlightLoadRecord.travel_date <= end_date)
            
        all_records = query.order_by(FlightLoadRecord.travel_date).all()
        
        inbound_records = []
        outbound_records = []
        
        for record in all_records:
            record_dict = record.to_dict()
            # Assuming flight_no '620' is inbound (ADD to KWI) and '621' is outbound (KWI to ADD)
            # This assumption is based on the process_flight_load_excel function logic
            if record.flight_no == '620':
                inbound_records.append(record_dict)
            elif record.flight_no == '621':
                outbound_records.append(record_dict)
                
        filtered_data = {
            'inbound': inbound_records,
            'outbound': outbound_records
        }
        
        # Apply flight type filter
        if flight_type == 'inbound':
            filtered_data['outbound'] = []
        elif flight_type == 'outbound':
            filtered_data['inbound'] = []
        
        if not filtered_data['inbound'] and not filtered_data['outbound']:
            return jsonify({'error': 'No flight load data available for the selected range'}), 404
        
        return jsonify({
            'success': True,
            'data': filtered_data
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@flight_load_bp.route('/summary')
def get_flight_load_summary():
    """Get summary statistics for flight load data"""
    # Check authentication
    is_admin = session.get('admin_logged_in', False)
    is_public = session.get('public_authenticated', False)
    
    if not is_admin and not is_public:
        return jsonify({'error': 'Authentication required'}), 401
    
    try:
        # Get filter parameters
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        
        query = FlightLoadRecord.query
        
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            query = query.filter(FlightLoadRecord.travel_date >= start_date)
        
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            query = query.filter(FlightLoadRecord.travel_date <= end_date)
            
        all_records = query.all()
        
        inbound = [r.to_dict() for r in all_records if r.flight_no == '620']
        outbound = [r.to_dict() for r in all_records if r.flight_no == '621']
        
        # Calculate summary statistics
        def calc_stats(records):
            if not records:
                return {
                    'avg_lf': 0.0,
                    'avg_lf_c': 0.0,
                    'avg_lf_y': 0.0,
                    'total_pax': 0,
                    'total_pax_c': 0,
                    'total_pax_y': 0,
                    'total_capacity': 0,
                    'flights_count': 0
                }
            
            # The records already contain the calculated LF values from the Excel, 
            # but the user complained about incorrect total passengers.
            # The most accurate way is to re-calculate LF from total pax and capacity.
            total_pax = sum(r['pax'] for r in records)
            total_pax_c = sum(r['pax_c'] for r in records)
            total_pax_y = sum(r['pax_y'] for r in records)
            total_capacity = sum(r['tot_cap'] for r in records)
            total_capacity_c = sum(r['c_cap'] for r in records)
            total_capacity_y = sum(r['y_cap'] for r in records)
            
            avg_lf = (total_pax / total_capacity) * 100 if total_capacity > 0 else 0.0
            avg_lf_c = (total_pax_c / total_capacity_c) * 100 if total_capacity_c > 0 else 0.0
            avg_lf_y = (total_pax_y / total_capacity_y) * 100 if total_capacity_y > 0 else 0.0
            
            return {
                'avg_lf': round(avg_lf, 2),
                'avg_lf_c': round(avg_lf_c, 2),
                'avg_lf_y': round(avg_lf_y, 2),
                'total_pax': total_pax,
                'total_pax_c': total_pax_c,
                'total_pax_y': total_pax_y,
                'total_capacity': total_capacity,
                'flights_count': len(records)
            }
        
        inbound_stats = calc_stats(inbound)
        outbound_stats = calc_stats(outbound)
        combined_stats = calc_stats(inbound + outbound)
        
        summary = {
            'inbound': inbound_stats,
            'outbound': outbound_stats,
            'combined': combined_stats,
            'total_passengers': inbound_stats['total_pax'] + outbound_stats['total_pax']
        }
        
        return jsonify({
            'success': True,
            'summary': summary
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
